import pandas as pd
import openpyxl
from datetime import datetime
from typing import Dict, List, Any
import logging

logger = logging.getLogger(__name__)

EXPECTED_SHEETS = {
    'Production': ['Date', 'Plant', 'Line', 'Cement_MT', 'Clinker_MT', 'Capacity_Util_%', 'Downtime_Hrs'],
    'Energy': ['Date', 'Plant', 'Power_kWh_Ton', 'Heat_kcal_kg', 'Fuel_Cost_Rs_Ton', 'AFR_%'],
    'Maintenance': ['Date', 'Plant', 'Equipment', 'Breakdown_Hrs', 'MTBF_Hrs', 'MTTR_Hrs'],
    'Quality': ['Date', 'Plant', 'Blaine', 'Strength_28D', 'Clinker_Factor'],
    'Sales_Logistics': ['Date', 'Plant', 'Region', 'Dispatch_MT', 'Realization_Rs_Ton', 'Freight_Rs_Ton', 'OTIF_%'],
    'Finance': ['Date', 'Plant', 'Cost_Rs_Ton', 'EBITDA_Rs_Ton', 'Margin_%']
}

class ExcelProcessor:
    def __init__(self, file_path: str):
        self.file_path = file_path
        self.workbook = openpyxl.load_workbook(file_path, data_only=True)
        self.validation_errors = []
        self.sheets_data = {}
    
    def validate_structure(self) -> Dict[str, Any]:
        """Validate Excel structure and return validation results"""
        results = {
            'status': 'ok',
            'errors': [],
            'warnings': [],
            'sheets_found': [],
            'mapping_suggestions': {}
        }
        
        available_sheets = self.workbook.sheetnames
        results['sheets_found'] = available_sheets
        
        # Check for expected sheets
        for expected_sheet in EXPECTED_SHEETS.keys():
            if expected_sheet not in available_sheets:
                results['warnings'].append(f"Sheet '{expected_sheet}' not found")
        
        return results
    
    def read_and_validate_data(self) -> Dict[str, pd.DataFrame]:
        """Read all sheets and validate data"""
        data = {}
        
        for sheet_name in EXPECTED_SHEETS.keys():
            if sheet_name in self.workbook.sheetnames:
                try:
                    df = pd.read_excel(self.file_path, sheet_name=sheet_name)
                    
                    # Normalize column names
                    df.columns = df.columns.str.strip()
                    
                    # Convert Date column
                    if 'Date' in df.columns:
                        df['Date'] = pd.to_datetime(df['Date'], errors='coerce')
                    
                    # Remove rows with missing dates
                    df = df.dropna(subset=['Date'])
                    
                    data[sheet_name] = df
                    logger.info(f"Sheet '{sheet_name}' loaded: {len(df)} rows")
                except Exception as e:
                    logger.error(f"Error reading sheet '{sheet_name}': {str(e)}")
                    self.validation_errors.append(f"Sheet '{sheet_name}': {str(e)}")
        
        return data
    
    def get_preview_data(self, max_rows: int = 10) -> Dict[str, List[Dict]]:
        """Get preview of data from each sheet"""
        preview = {}
        data = self.read_and_validate_data()
        
        for sheet_name, df in data.items():
            preview[sheet_name] = df.head(max_rows).to_dict('records')
            # Convert datetime to string for JSON serialization
            for row in preview[sheet_name]:
                for key, value in row.items():
                    if pd.isna(value):
                        row[key] = None
                    elif isinstance(value, pd.Timestamp):
                        row[key] = value.strftime('%Y-%m-%d')
                    elif isinstance(value, (pd.Int64Dtype, int)):
                        row[key] = int(value)
                    elif isinstance(value, float):
                        row[key] = float(value)
        
        return preview
    
    def get_stats(self) -> Dict[str, Any]:
        """Get statistics about the data"""
        data = self.read_and_validate_data()
        stats = {
            'rowsPerSheet': {},
            'dateRange': {},
            'plants': []
        }
        
        all_plants = set()
        
        for sheet_name, df in data.items():
            stats['rowsPerSheet'][sheet_name] = len(df)
            
            if 'Date' in df.columns and len(df) > 0:
                stats['dateRange'][sheet_name] = {
                    'start': df['Date'].min().strftime('%Y-%m-%d'),
                    'end': df['Date'].max().strftime('%Y-%m-%d')
                }
            
            if 'Plant' in df.columns:
                all_plants.update(df['Plant'].dropna().unique().tolist())
        
        stats['plants'] = sorted(list(all_plants))
        
        return stats
